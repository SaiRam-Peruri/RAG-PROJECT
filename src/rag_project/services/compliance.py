"""
Compliance matrix service — extracts requirements and builds tracking matrix.
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..config import settings
from ..core.chroma_client import get_chroma_manager
from ..core.validation import sanitize_opportunity_name
from ..logging_config import get_logger

logger = get_logger("compliance")


def extract_requirements(text: str) -> List[Dict]:
    """
    Extract requirement statements from RFP text.
    Looks for: shall, must, will, should.
    """
    requirements: List[Dict] = []
    sentences = re.split(r'(?<=[.!?])\s+', text)
    req_id = 1

    for sentence in sentences:
        if len(sentence) < 20:
            continue

        sentence = sentence.strip()
        req_type = None

        if re.search(r'\b(shall|must)\b', sentence, re.IGNORECASE):
            req_type = "Mandatory"
        elif re.search(r'\bwill\b', sentence, re.IGNORECASE):
            req_type = "Statement of Work"
        elif re.search(r'\bshould\b', sentence, re.IGNORECASE):
            req_type = "Desirable"

        if req_type:
            requirements.append({
                "req_id": f"REQ-{req_id:04d}",
                "type": req_type,
                "text": sentence[:500],
                "section": "Unknown",
                "page": None,
            })
            req_id += 1

    return requirements


def categorize_requirements(requirements: List[Dict]) -> Dict[str, List[Dict]]:
    """Categorize requirements by domain."""
    categories: Dict[str, List[Dict]] = {
        "Technical": [], "Management": [], "Deliverables": [],
        "Reporting": [], "Personnel": [], "Security": [], "Other": [],
    }

    keyword_map = {
        "Technical": ("technical", "system", "software", "hardware", "architecture", "cloud", "api"),
        "Management": ("manage", "plan", "organization", "quality", "process"),
        "Deliverables": ("deliver", "submit", "provide", "produce"),
        "Reporting": ("report", "status", "meeting", "communication"),
        "Personnel": ("personnel", "staff", "key person", "resume", "clearance"),
        "Security": ("security", "classified", "fisma", "ato", "encryption"),
    }

    for req in requirements:
        text_lower = req["text"].lower()
        categorized = False

        for category, keywords in keyword_map.items():
            if any(kw in text_lower for kw in keywords):
                categories[category].append(req)
                categorized = True
                break

        if not categorized:
            categories["Other"].append(req)

    return categories


def generate_compliance_matrix(
    opportunity: str,
    output_path: Path | None = None,
) -> Dict:
    """
    Generate a compliance matrix for an opportunity.

    Returns:
        {"requirements": int, "mandatory": int, "desirable": int, "output_file": str}
    """
    opportunity = sanitize_opportunity_name(opportunity)
    settings.require_api_key()

    manager = get_chroma_manager()
    coll = manager.get_auth_collection()

    # Query for requirements
    where = {
        "$and": [
            {"opportunity": opportunity},
            {"stage": {"$ne": "award"}},
            {"doc_role": {"$in": ["technical_requirements", "instructions", "general", "amendment_qa"]}},
        ]
    }

    result = coll.query(
        query_texts=["requirements specifications shall must will should deliverables"],
        n_results=50,
        where=where,
    )

    docs = result["documents"][0]
    metas = result["metadatas"][0]

    # Extract requirements
    all_requirements: List[Dict] = []
    for doc, meta in zip(docs, metas):
        reqs = extract_requirements(doc)
        for req in reqs:
            req["section"] = meta.get("filename", "Unknown")
            req["page"] = meta.get("page")
        all_requirements.extend(reqs)

    logger.info("Extracted %d requirements for %s", len(all_requirements), opportunity)

    # Create Excel output
    if output_path is None:
        output_path = settings.project_root / f"{opportunity}_Compliance_Matrix.xlsx"

    _create_excel_matrix(all_requirements, opportunity, output_path)

    stats = {
        "requirements": len(all_requirements),
        "mandatory": sum(1 for r in all_requirements if r["type"] == "Mandatory"),
        "desirable": sum(1 for r in all_requirements if r["type"] == "Desirable"),
        "sow": sum(1 for r in all_requirements if r["type"] == "Statement of Work"),
        "output_file": str(output_path),
    }

    logger.info("Compliance matrix saved: %s", output_path)
    return stats


def _create_excel_matrix(requirements: List[Dict], opportunity: str, output_path: Path):
    """Create Excel compliance matrix with proper formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compliance Matrix"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "Req ID", "Category", "Type", "Requirement Text",
        "Source (Section/Page)", "Response Location",
        "Compliance Status", "Notes",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    widths = [12, 15, 12, 60, 20, 25, 15, 30]
    for i, w in enumerate(widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w

    categorized = categorize_requirements(requirements)
    row = 2

    for category, reqs in categorized.items():
        for req in reqs:
            ws.cell(row=row, column=1).value = req["req_id"]
            ws.cell(row=row, column=2).value = category
            ws.cell(row=row, column=3).value = req["type"]
            ws.cell(row=row, column=4).value = req["text"]
            ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
            source = req.get("section", "Unknown")
            if req.get("page"):
                source += f" p.{req['page']}"
            ws.cell(row=row, column=5).value = source
            ws.cell(row=row, column=6).value = ""
            ws.cell(row=row, column=7).value = "Not Started"
            ws.cell(row=row, column=8).value = ""
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border
            row += 1

    # Metadata sheet
    meta_ws = wb.create_sheet("Metadata")
    meta_ws["A1"], meta_ws["B1"] = "Opportunity:", opportunity
    meta_ws["A2"], meta_ws["B2"] = "Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    meta_ws["A3"], meta_ws["B3"] = "Total Requirements:", len(requirements)

    wb.save(str(output_path))
