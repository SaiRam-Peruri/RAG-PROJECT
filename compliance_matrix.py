"""
Compliance Matrix Generator
Extracts requirements from RFP and builds a compliance tracking matrix.

Parses "shall", "must", "will", "should" statements and creates Excel matrix.
"""

import os
import re
from typing import List, Dict, Tuple
from pathlib import Path
from datetime import datetime

import chromadb
from chromadb.utils import embedding_functions
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def extract_requirements(text: str) -> List[Dict]:
    """
    Extract requirement statements from RFP text.
    
    Looks for sentences containing: shall, must, will, should
    """
    requirements = []
    
    # Split into sentences (simple approach)
    sentences = re.split(r'(?<=[.!?])\s+', text)
    
    req_id = 1
    for sentence in sentences:
        # Skip very short sentences
        if len(sentence) < 20:
            continue
        
        sentence = sentence.strip()
        
        # Identify requirement type
        req_type = None
        if re.search(r'\b(shall|must)\b', sentence, re.IGNORECASE):
            req_type = "Mandatory"
        elif re.search(r'\bwill\b', sentence, re.IGNORECASE):
            req_type = "Statement of Work"
        elif re.search(r'\bshould\b', sentence, re.IGNORECASE):
            req_type = "Desirable"
        
        if req_type:
            requirements.append({
                'req_id': f"REQ-{req_id:04d}",
                'type': req_type,
                'text': sentence[:500],  # Limit length
                'section': 'Unknown',  # Will be populated from metadata
                'page': None
            })
            req_id += 1
    
    return requirements


def categorize_requirements(requirements: List[Dict]) -> Dict[str, List[Dict]]:
    """Categorize requirements by type for better organization."""
    categories = {
        'Technical': [],
        'Management': [],
        'Deliverables': [],
        'Reporting': [],
        'Personnel': [],
        'Security': [],
        'Other': []
    }
    
    for req in requirements:
        text_lower = req['text'].lower()
        
        if any(kw in text_lower for kw in ['technical', 'system', 'software', 'hardware', 'architecture', 'cloud', 'api']):
            categories['Technical'].append(req)
        elif any(kw in text_lower for kw in ['manage', 'plan', 'organization', 'quality', 'process']):
            categories['Management'].append(req)
        elif any(kw in text_lower for kw in ['deliver', 'submit', 'provide', 'produce']):
            categories['Deliverables'].append(req)
        elif any(kw in text_lower for kw in ['report', 'status', 'meeting', 'communication']):
            categories['Reporting'].append(req)
        elif any(kw in text_lower for kw in ['personnel', 'staff', 'key person', 'resume', 'clearance']):
            categories['Personnel'].append(req)
        elif any(kw in text_lower for kw in ['security', 'classified', 'fisma', 'ato', 'encryption']):
            categories['Security'].append(req)
        else:
            categories['Other'].append(req)
    
    return categories


def create_compliance_matrix(requirements: List[Dict], opportunity: str, output_path: Path):
    """
    Create Excel compliance matrix.
    
    Columns:
    - Req ID
    - Category
    - Requirement Type (Mandatory/Desirable)
    - Requirement Text
    - Source (Section/Page)
    - Proposal Response Location
    - Compliance Status
    - Notes
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compliance Matrix"
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Req ID', 'Category', 'Type', 'Requirement Text', 
        'Source (Section/Page)', 'Response Location', 
        'Compliance Status', 'Notes'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 12  # Req ID
    ws.column_dimensions['B'].width = 15  # Category
    ws.column_dimensions['C'].width = 12  # Type
    ws.column_dimensions['D'].width = 60  # Requirement Text
    ws.column_dimensions['E'].width = 20  # Source
    ws.column_dimensions['F'].width = 25  # Response Location
    ws.column_dimensions['G'].width = 15  # Compliance Status
    ws.column_dimensions['H'].width = 30  # Notes
    
    # Categorize requirements
    categorized = categorize_requirements(requirements)
    
    # Populate rows
    row = 2
    for category, reqs in categorized.items():
        if not reqs:
            continue
        
        for req in reqs:
            ws.cell(row=row, column=1).value = req['req_id']
            ws.cell(row=row, column=2).value = category
            ws.cell(row=row, column=3).value = req['type']
            ws.cell(row=row, column=4).value = req['text']
            ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical='top')
            
            source = f"{req.get('section', 'Unknown')}"
            if req.get('page'):
                source += f" p.{req['page']}"
            ws.cell(row=row, column=5).value = source
            
            # Empty cells for manual population
            ws.cell(row=row, column=6).value = ""  # Response Location
            ws.cell(row=row, column=7).value = "Not Started"  # Compliance Status
            ws.cell(row=row, column=8).value = ""  # Notes
            
            # Apply borders
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border
            
            row += 1
    
    # Add metadata sheet
    meta_ws = wb.create_sheet("Metadata")
    meta_ws['A1'] = "Opportunity:"
    meta_ws['B1'] = opportunity
    meta_ws['A2'] = "Generated:"
    meta_ws['B2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    meta_ws['A3'] = "Total Requirements:"
    meta_ws['B3'] = len(requirements)
    meta_ws['A4'] = "Mandatory:"
    meta_ws['B4'] = sum(1 for r in requirements if r['type'] == 'Mandatory')
    meta_ws['A5'] = "Desirable:"
    meta_ws['B5'] = sum(1 for r in requirements if r['type'] == 'Desirable')
    
    # Add instructions sheet
    inst_ws = wb.create_sheet("Instructions")
    instructions = [
        "HOW TO USE THIS COMPLIANCE MATRIX",
        "",
        "1. RESPONSE LOCATION: Enter where in your proposal you address each requirement",
        "   Example: 'Section 3.2.1, p.15' or 'Technical Approach Volume, Section 2.3'",
        "",
        "2. COMPLIANCE STATUS: Update as you draft proposal",
        "   Options: Not Started | In Progress | Completed | N/A",
        "",
        "3. NOTES: Track any clarifications, assumptions, or issues",
        "   Example: 'Requested clarification via Q&A' or 'Addressed by subcontractor XYZ'",
        "",
        "4. REVIEW PROCESS:",
        "   - Technical team: Review Technical & Security requirements",
        "   - Management team: Review Management & Personnel requirements",
        "   - Contracts: Review all Deliverables & Reporting requirements",
        "",
        "5. BEFORE SUBMISSION:",
        "   - Verify ALL mandatory requirements have Compliance Status = 'Completed'",
        "   - Ensure Response Location is populated for every requirement",
        "   - Color-code: Green = Completed, Yellow = In Progress, Red = Not Started",
        "",
        "NOTE: This matrix was AUTO-GENERATED. Review all requirements for accuracy."
    ]
    
    for i, line in enumerate(instructions, 1):
        inst_ws.cell(row=i, column=1).value = line
        if i == 1:
            inst_ws.cell(row=i, column=1).font = Font(bold=True, size=14)
    
    inst_ws.column_dimensions['A'].width = 100
    
    # Save
    wb.save(str(output_path))


def main():
    if not os.getenv("OPENAI_API_KEY"):
        raise SystemExit("OPENAI_API_KEY not set.\n")
    
    print("\n=== Compliance Matrix Generator ===\n")
    
    # Setup
    client_chroma = chromadb.PersistentClient(path="chroma_db")
    embedder = embedding_functions.OpenAIEmbeddingFunction(
        api_key=os.environ["OPENAI_API_KEY"],
        model_name="text-embedding-3-large"
    )
    coll = client_chroma.get_collection("authoritative", embedding_function=embedder)
    
    # Get available opportunities
    result = coll.get()
    opportunities = sorted(set(m.get('opportunity', 'unknown') for m in result['metadatas'] if m.get('opportunity') != 'unknown'))
    
    print("Available opportunities:")
    for i, opp in enumerate(opportunities, 1):
        print(f"  {i}. {opp}")
    
    choice = input("\nSelect opportunity (1-N): ").strip()
    opportunity = opportunities[int(choice) - 1]
    
    print(f"\n→ Extracting requirements from {opportunity}...")
    
    # Query for RFP documents (exclude awards, include requirements)
    where = {
        "$and": [
            {"opportunity": opportunity},
            {"stage": {"$ne": "award"}},
            {"doc_role": {"$in": ["technical_requirements", "instructions", "general", "amendment_qa"]}}
        ]
    }
    
    query_result = coll.query(
        query_texts=["requirements specifications shall must will should deliverables"],
        n_results=50,  # Get lots of content
        where=where
    )
    
    docs = query_result["documents"][0]
    metas = query_result["metadatas"][0]
    
    print(f"  Found {len(docs)} relevant document chunks")
    
    # Extract requirements from all chunks
    all_requirements = []
    for doc, meta in zip(docs, metas):
        requirements = extract_requirements(doc)
        
        # Add metadata to requirements
        for req in requirements:
            req['section'] = meta.get('filename', 'Unknown')
            req['page'] = meta.get('page')
        
        all_requirements.extend(requirements)
    
    print(f"  Extracted {len(all_requirements)} requirement statements")
    
    # Create compliance matrix
    output_file = Path(f"{opportunity}_Compliance_Matrix.xlsx")
    print(f"\n→ Creating compliance matrix: {output_file}")
    
    create_compliance_matrix(all_requirements, opportunity, output_file)
    
    print(f"\n✅ Compliance matrix created!\n")
    print(f"Output: {output_file}")
    print(f"Total requirements: {len(all_requirements)}")
    print(f"  - Mandatory: {sum(1 for r in all_requirements if r['type'] == 'Mandatory')}")
    print(f"  - Desirable: {sum(1 for r in all_requirements if r['type'] == 'Desirable')}")
    print(f"  - SOW: {sum(1 for r in all_requirements if r['type'] == 'Statement of Work')}")
    
    # Show categorization breakdown
    categorized = categorize_requirements(all_requirements)
    print("\nRequirements by category:")
    for category, reqs in categorized.items():
        if reqs:
            print(f"  - {category}: {len(reqs)}")
    
    print("\n💡 Next steps:")
    print("  1. Open the Excel file")
    print("  2. Review requirements for accuracy")
    print("  3. Assign to proposal sections")
    print("  4. Track compliance as you draft")


if __name__ == "__main__":
    main()
